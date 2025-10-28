import openpyxl
import re
from io import BytesIO

def normalize_key(text):
    """Normaliza texto a un formato de clave (snake_case)."""
    if not isinstance(text, str):
        text = str(text)
    text = text.lower()
    replacements = {
        " ": "_", "º": "nro", ".": "", ":": "", "ñ": "n", "ó": "o", "ö": "o",
        "é": "e", "í": "i", "á": "a", "ú": "u", "ü": "u", "-": "_", "(": "", ")": "",
        "/": "_"
    }
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    text = re.sub(r'_+', '_', text)
    return text.strip("_")

def procesar_excel(uploaded_file):
    """
    Lee un archivo Excel TMERT, extrae datos generales y filtra casos ART
    con riesgo "Intermedio" utilizando toda la lógica de validación, incluyendo
    la condición de "No Aceptable" como pre-filtro.
    """
    if not uploaded_file:
        return None, []

    try:
        content = uploaded_file.getvalue()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception:
        return None, []

    datos_generales = {"antecedentes_empresa": {}, "centro_trabajo": {}, "responsable_protocolo": {}}
    casos_art_intermedio = []

    # DEBUG: Imprimir nombres de hojas disponibles
    print(f"DEBUG - Hojas disponibles en Excel: {wb.sheetnames}")

    # --- PASO 1: Construir lista de puestos válidos desde la Hoja '2' ---
    valid_job_numbers = set()
    puestos_detalle = {}  # Nuevo: Guardar información detallada de cada puesto
    
    try:
        hoja2 = wb["2"]
        
        # DEBUG: Mostrar encabezados de Hoja 2
        print("DEBUG - Explorando estructura de Hoja 2:")
        for col_idx in range(1, 20):  # Explorar más columnas
            header = hoja2.cell(row=12, column=col_idx).value  # Fila 12 suele tener encabezados en Hoja 2
            print(f"  Columna {col_idx} ({chr(64+col_idx)}): '{header}'")
        
        for fila_idx in range(13, 200): 
            nro_puesto = str(hoja2.cell(row=fila_idx, column=2).value or "").strip()
            area = str(hoja2.cell(row=fila_idx, column=3).value or "").strip()
            puesto = str(hoja2.cell(row=fila_idx, column=4).value or "").strip()
            
            if nro_puesto and nro_puesto != "0" and area and area != "0" and puesto and puesto != "0":
                valid_job_numbers.add(nro_puesto)
                
                # Extraer información adicional de la Hoja 2 (columnas correctas)
                puestos_detalle[nro_puesto] = {
                    "descripcion_de_la_tarea": str(hoja2.cell(row=fila_idx, column=6).value or "").strip(),  # Columna F
                    "horario_de_funcionamiento_detalle": str(hoja2.cell(row=fila_idx, column=7).value or "").strip(),  # Columna G
                    "hhex_dia_detalle": str(hoja2.cell(row=fila_idx, column=8).value or "").strip(),  # Columna H
                    "tipo_contrato": str(hoja2.cell(row=fila_idx, column=12).value or "").strip(),  # Columna L
                    "tipo_remuneracion": str(hoja2.cell(row=fila_idx, column=13).value or "").strip(),  # Columna M
                    "duracion_min": str(hoja2.cell(row=fila_idx, column=14).value or "").strip(),  # Columna N
                    "pausas": str(hoja2.cell(row=fila_idx, column=15).value or "").strip(),  # Columna O
                    "rotacion_tarea": str(hoja2.cell(row=fila_idx, column=16).value or "").strip(),  # Columna P
                    "equipos_herramientas": str(hoja2.cell(row=fila_idx, column=17).value or "").strip(),  # Columna Q
                    "caracteristicas_ambientes_espacios_trabajo": str(hoja2.cell(row=fila_idx, column=18).value or "").strip(),  # Columna R
                    "caracteristicas_disposicion_espacial_puesto": str(hoja2.cell(row=fila_idx, column=19).value or "").strip(),  # Columna S
                }
                
    except KeyError as e:
        print(f"DEBUG - Error al leer Hoja 2: {e}")
        pass

    # --- PASO 2: Extraer Datos Generales (Hoja '1') ---
    try:
        hoja1 = wb["1"]
        mapeo_hoja1 = {
            "antecedentes_empresa": {
                "Razón Social": (15, 'E'), 
                "RUT Empresa": (15, 'L'), 
                "Actividad Económica": (17, 'E'), 
                "Código CIIU": (17, 'L'),  # Agregar código CIIU
                "Dirección": (19, 'E'), 
                "Comuna": (19, 'L'), 
                "Nombre Representante Legal": (21, 'E')
            },
            "centro_trabajo": {
                "Nombre del centro de trabajo": (27, 'E'), 
                "Dirección": (29, 'E'), 
                "Comuna": (29, 'L'),
                "Total trabajadores CT": (31, 'L'),  # Cambiar a columna L
                "Centro trabajo": (27, 'E'),  # Agregar alias para centro_trabajo
                "Nombre": (27, 'E')  # Otro alias posible
            },
            "responsable_protocolo": {
                "Nombre responsable": (35, 'E'), 
                "Cargo": (37, 'E')
            }
        }
        for seccion, campos in mapeo_hoja1.items():
            for etiqueta, (fila, col) in campos.items():
                valor = hoja1[f"{col}{fila}"].value
                if valor:
                    datos_generales[seccion][normalize_key(etiqueta)] = str(valor).strip()
        
        # DEBUG: Imprimir datos extraídos
        print(f"DEBUG - Datos generales extraídos:")
        for seccion, datos in datos_generales.items():
            print(f"  {seccion}: {datos}")
            
    except KeyError as e:
        print(f"DEBUG - Error al leer Hoja 1: {e}")
        pass

    # --- PASO 3: Extraer y Filtrar Casos ART (Hoja '4') con LÓGICA FINAL ---
    try:
        hoja4 = wb["4"]
        
        # DEBUG: Mostrar encabezados de columnas para entender la estructura
        print("DEBUG - Explorando estructura de Hoja 4:")
        for col_idx in range(1, 25):  # Columnas A-X
            header = hoja4.cell(row=13, column=col_idx).value  # Fila 13 suele tener encabezados
            print(f"  Columna {col_idx} ({chr(64+col_idx)}): '{header}'")
        
        for fila_idx in range(14, 200):
            nro_puesto = str(hoja4.cell(row=fila_idx, column=2).value or "").strip()
            area = str(hoja4.cell(row=fila_idx, column=3).value or "").strip()
            puesto = str(hoja4.cell(row=fila_idx, column=4).value or "").strip()

            # Filtro 1: Consistencia interna de la fila en Hoja 4
            if not (nro_puesto and nro_puesto != "0" and area and area != "0" and puesto and puesto != "0"):
                continue

            # Filtro 2: Validación cruzada con la Hoja 2
            if nro_puesto not in valid_job_numbers:
                continue

            # Filtro 3 (CRUCIAL): Pre-filtro de Condición Aceptable (Columna Q)
            condicion_aceptable_val = str(hoja4.cell(row=fila_idx, column=17).value or "").lower()
            if "no aceptable" not in condicion_aceptable_val:
                continue

            # Filtro 4: Nivel de Riesgo Intermedio (Columna X)
            riesgo_val = str(hoja4.cell(row=fila_idx, column=24).value or "").lower()
            if "intermedio" in riesgo_val or "no crítico" in riesgo_val.replace("í", "i"):
                # Datos básicos de Hoja 4
                caso = {
                    # Datos básicos
                    "nro": nro_puesto,
                    "area": area,
                    "puesto": puesto,
                    "tarea": str(hoja4.cell(row=fila_idx, column=5).value or "").strip(),
                    
                    # Mapeo para plantilla (nombres exactos)
                    "area_de_trabajo": area,
                    "puesto_de_trabajo": puesto,
                    "tareas_del_puesto": str(hoja4.cell(row=fila_idx, column=5).value or "").strip(),
                    
                    # Trabajadores expuestos (Columnas I y J)
                    "n_trab_exp_hombre": str(hoja4.cell(row=fila_idx, column=9).value or "0").strip(),
                    "n_trab_exp_mujer": str(hoja4.cell(row=fila_idx, column=10).value or "0").strip(),
                    "n°_trab_exp_hombre": str(hoja4.cell(row=fila_idx, column=9).value or "0").strip(),
                    "n°_trab_exp_mujer": str(hoja4.cell(row=fila_idx, column=10).value or "0").strip(),
                    
                    # Datos de horario de Hoja 4 (Columna F)
                    "horario_de_funcionamiento": str(hoja4.cell(row=fila_idx, column=6).value or "").strip(),
                    "hhex_dia": str(hoja4.cell(row=fila_idx, column=7).value or "").strip(),
                }
                
                # Combinar con datos detallados de Hoja 2
                if nro_puesto in puestos_detalle:
                    # Combinar manteniendo los datos de Hoja 4 cuando corresponde
                    detalle = puestos_detalle[nro_puesto]
                    caso.update({
                        # Mantener descripción de tarea si está más completa en Hoja 2
                        "descripcion_de_la_tarea": detalle.get("descripcion_de_la_tarea", caso["tarea"]),
                        "equipos_herramientas": detalle.get("equipos_herramientas", "No especificado"),
                        "caracteristicas_ambientes_espacios_trabajo": detalle.get("caracteristicas_ambientes_espacios_trabajo", "No especificado"),
                        "caracteristicas_disposicion_espacial_puesto": detalle.get("caracteristicas_disposicion_espacial_puesto", "No especificado"),
                        "tipo_remuneracion": detalle.get("tipo_remuneracion", "No especificado"),
                        "duracion_min": detalle.get("duracion_min", "No especificado"),
                        "pausas": detalle.get("pausas", "No especificado"),
                        "tipo_contrato": detalle.get("tipo_contrato", "No especificado"),
                        "rotacion_tarea": detalle.get("rotacion_tarea", "No especificado"),
                    })
                else:
                    # Valores por defecto si no hay datos en Hoja 2
                    caso.update({
                        "descripcion_de_la_tarea": caso["tarea"],  # Usar tarea como descripción
                        "equipos_herramientas": "No especificado",
                        "caracteristicas_ambientes_espacios_trabajo": "No especificado",
                        "caracteristicas_disposicion_espacial_puesto": "No especificado",
                        "tipo_remuneracion": "No especificado",
                        "duracion_min": "No especificado",
                        "pausas": "No especificado",
                        "tipo_contrato": "No especificado",
                        "rotacion_tarea": "No especificado",
                    })
                
                # Calcular total trabajadores expuestos
                try:
                    total_hombres = int(caso["n_trab_exp_hombre"]) if caso["n_trab_exp_hombre"].isdigit() else 0
                    total_mujeres = int(caso["n_trab_exp_mujer"]) if caso["n_trab_exp_mujer"].isdigit() else 0
                    caso["total_trabajadores_expuestos_puesto"] = str(total_hombres + total_mujeres)
                except:
                    caso["total_trabajadores_expuestos_puesto"] = "0"
                    
                casos_art_intermedio.append(caso)
    
    except KeyError as e:
        print(f"DEBUG - Error al leer Hoja 4: {e}")
        return datos_generales, []

    print(f"DEBUG - Total casos ART extraídos: {len(casos_art_intermedio)}")
    if casos_art_intermedio:
        print(f"DEBUG - Primer caso como ejemplo:")
        for key, value in casos_art_intermedio[0].items():
            print(f"  {key}: '{value}'")

    return datos_generales, casos_art_intermedio
